"""Quick Entry screen for StockMate.

Provides a fast data-entry interface for adding inventory items one-at-a-time
or in batch-scan mode. Supports auto-fetch of device details, lock toggles
for repeated fields, autocomplete, smart focus navigation, and duplicate
checking against the current inventory.
"""

from __future__ import annotations

import datetime
import logging
import os
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from typing import Any

import pandas as pd

from core.scraper import PhoneScraper
from gui.base import AutocompleteEntry, BaseScreen

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Default pick-lists — used when config has no custom values
# ---------------------------------------------------------------------------

_DEFAULT_COLORS: list[str] = [
    "Black",
    "White",
    "Blue",
    "Green",
    "Red",
    "Gold",
    "Silver",
    "Gray",
    "Purple",
    "Pink",
    "Orange",
    "Brown",
]

_DEFAULT_GRADES: list[str] = ["A+", "A", "B+", "B", "C", "New"]

_DEFAULT_CONDITIONS: list[str] = [
    "Like New",
    "Good",
    "Fair",
    "Poor",
    "Scratched",
    "Dented",
]


class QuickEntryScreen(BaseScreen):
    """Fast inventory entry screen with single and batch modes."""

    def __init__(self, parent: tk.Misc, app_context: dict[str, Any]) -> None:
        super().__init__(parent, app_context)
        self.scraper = PhoneScraper()

        # -- State variables ------------------------------------------------
        self.var_imei = tk.StringVar()
        self.var_model = tk.StringVar()
        self.var_ram_rom = tk.StringVar()
        self.var_price = tk.StringVar()
        self.var_color = tk.StringVar()
        self.var_supplier = tk.StringVar()
        self.var_grade = tk.StringVar()
        self.var_condition = tk.StringVar()

        # Toggles / Locks
        self.var_lock_supplier = tk.BooleanVar(value=False)
        self.var_lock_grade = tk.BooleanVar(value=False)
        self.var_lock_color = tk.BooleanVar(value=False)
        self.var_auto_fetch = tk.BooleanVar(value=True)
        self.var_manual_model = tk.BooleanVar(value=False)
        self.var_print_after_save = tk.BooleanVar(value=False)
        self.var_batch_mode = tk.BooleanVar(value=False)

        self.target_file_key = tk.StringVar()
        self.file_options: list[str] = []
        self.file_display_map: dict[str, str] = {}  # Display Name -> Full Path
        self.batch_queue: list[str] = []
        self.fetched_data: dict[str, dict[str, Any]] = {}

        self._init_ui()

        # Shortcuts
        self.bind_all("<Control-Return>", lambda e: self._focus_batch_start(), add="+")

    # -- lifecycle -----------------------------------------------------------

    def focus_primary(self) -> None:
        self.ent_imei.focus_set()

    def on_show(self) -> None:
        self._refresh_files()
        self.ent_imei.focus_set()

    # -- UI ------------------------------------------------------------------

    def _init_ui(self) -> None:
        # Header
        self.add_header("Quick Entry", help_section="Core Features")

        # --- Top Bar: File & Settings ---
        top_bar = ttk.Frame(self, padding=10)
        top_bar.pack(fill=tk.X)

        ttk.Label(top_bar, text="Target File:", font=("bold",)).pack(side=tk.LEFT)
        self.combo_file = ttk.Combobox(
            top_bar,
            textvariable=self.target_file_key,
            state="readonly",
            width=35,
        )
        self.combo_file.pack(side=tk.LEFT, padx=5)

        ttk.Button(top_bar, text="+ New", command=self._create_new_file, width=6).pack(
            side=tk.LEFT,
        )
        ttk.Button(top_bar, text="Refresh", command=self._refresh_files).pack(
            side=tk.LEFT,
            padx=5,
        )

        ttk.Checkbutton(
            top_bar,
            text="Auto-Fetch (Bg)",
            variable=self.var_auto_fetch,
        ).pack(side=tk.RIGHT, padx=10)
        ttk.Checkbutton(
            top_bar,
            text="Manual Model",
            variable=self.var_manual_model,
            command=self._toggle_model_entry,
        ).pack(side=tk.RIGHT, padx=10)
        ttk.Checkbutton(
            top_bar,
            text="Batch Mode",
            variable=self.var_batch_mode,
            command=self._toggle_batch_mode,
            bootstyle="round-toggle",
        ).pack(side=tk.RIGHT, padx=10)

        # --- Main Form (Single Entry) ---
        self.form_frame = ttk.LabelFrame(self, text="New Entry Details", padding=15)

        # --- Batch Frame (Hidden by default) ---
        self.batch_container = ttk.Frame(self)

        # Grid Layout Helper
        r = 0

        def add_field(
            label: str,
            var: tk.StringVar,
            widget_cls: type[ttk.Entry] = ttk.Entry,
            **kwargs: Any,
        ) -> tuple[Any, int]:
            nonlocal r
            ttk.Label(self.form_frame, text=label).grid(
                row=r,
                column=0,
                sticky=tk.W,
                pady=5,
            )
            w = widget_cls(self.form_frame, textvariable=var, **kwargs)
            w.grid(row=r, column=1, sticky=tk.EW, padx=5, pady=5)
            w.bind("<Return>", lambda e: w.tk_focusNext().focus())
            return w, r

        # 1. IMEI
        ttk.Label(self.form_frame, text="IMEI / SN:").grid(
            row=r,
            column=0,
            sticky=tk.W,
            pady=5,
        )
        self.ent_imei = ttk.Entry(
            self.form_frame,
            textvariable=self.var_imei,
            width=30,
            font=("Consolas", 11),
        )
        self.ent_imei.grid(row=r, column=1, sticky=tk.EW, padx=5, pady=5)
        self.ent_imei.bind("<Return>", self._on_imei_enter)
        ttk.Button(
            self.form_frame,
            text="Fetch Now",
            command=lambda: self._fetch_info_bg(self.var_imei.get()),
        ).grid(row=r, column=2, padx=5)
        r += 1

        # 2. Model
        ttk.Label(self.form_frame, text="Model Name:").grid(
            row=r,
            column=0,
            sticky=tk.W,
            pady=5,
        )

        models_list = self._get_unique_column_values("model")
        self.ent_model = AutocompleteEntry(
            self.form_frame,
            completion_list=models_list,
            textvariable=self.var_model,
            state="readonly",
        )
        self.ent_model.grid(row=r, column=1, sticky=tk.EW, padx=5, pady=5)
        self.ent_model.bind("<Return>", lambda e: self.ent_ram_rom.focus_set())
        r += 1

        # 3. Spec Row (RAM/ROM / Color)
        f_specs = ttk.Frame(self.form_frame)
        f_specs.grid(row=r, column=0, columnspan=3, sticky=tk.EW, pady=5)

        ttk.Label(f_specs, text="RAM/ROM:").pack(side=tk.LEFT)

        specs_list = self._get_unique_column_values("ram_rom")
        self.ent_ram_rom = AutocompleteEntry(
            f_specs,
            completion_list=specs_list,
            textvariable=self.var_ram_rom,
            width=15,
        )
        self.ent_ram_rom.pack(side=tk.LEFT, padx=5)
        self.ent_ram_rom.bind("<Return>", lambda e: self.cb_col.focus_set())

        ttk.Label(f_specs, text="Color:").pack(side=tk.LEFT, padx=(10, 0))
        colors = self._config_list("colors", _DEFAULT_COLORS)
        self.cb_col = ttk.Combobox(
            f_specs,
            textvariable=self.var_color,
            values=colors,
            width=15,
        )
        self.cb_col.pack(side=tk.LEFT, padx=5)
        self.cb_col.bind("<Return>", lambda e: self.ent_price.focus_set())
        ttk.Checkbutton(f_specs, text="\U0001f512", variable=self.var_lock_color).pack(
            side=tk.LEFT,
        )
        r += 1

        # 4. Price & Supplier
        f_ps = ttk.Frame(self.form_frame)
        f_ps.grid(row=r, column=0, columnspan=3, sticky=tk.EW, pady=5)

        ttk.Label(f_ps, text="Price:").pack(side=tk.LEFT)
        self.ent_price = ttk.Entry(f_ps, textvariable=self.var_price, width=10)
        self.ent_price.pack(side=tk.LEFT, padx=5)
        self.ent_price.bind("<Return>", lambda e: self.ent_supplier.focus_set())

        ttk.Label(f_ps, text="Supplier:").pack(side=tk.LEFT, padx=(15, 0))
        supp_list = self._get_unique_column_values("supplier")
        self.ent_supplier = AutocompleteEntry(
            f_ps,
            completion_list=supp_list,
            textvariable=self.var_supplier,
            width=15,
        )
        self.ent_supplier.pack(side=tk.LEFT, padx=5)
        ttk.Checkbutton(f_ps, text="\U0001f512", variable=self.var_lock_supplier).pack(
            side=tk.LEFT,
        )
        self.ent_supplier.bind("<Return>", lambda e: self.cb_grade.focus_set())
        r += 1

        # 5. Grade & Condition
        f_gc = ttk.Frame(self.form_frame)
        f_gc.grid(row=r, column=0, columnspan=3, sticky=tk.EW, pady=5)

        ttk.Label(f_gc, text="Grade:").pack(side=tk.LEFT)
        grades = self._config_list("grades", _DEFAULT_GRADES)
        self.cb_grade = AutocompleteEntry(
            f_gc,
            completion_list=grades,
            textvariable=self.var_grade,
            width=5,
        )
        self.cb_grade.pack(side=tk.LEFT, padx=5)
        ttk.Checkbutton(f_gc, text="\U0001f512", variable=self.var_lock_grade).pack(
            side=tk.LEFT,
        )

        ttk.Label(f_gc, text="Condition:").pack(side=tk.LEFT, padx=(15, 0))
        conditions = self._config_list("conditions", _DEFAULT_CONDITIONS)
        self.ent_cond = AutocompleteEntry(
            f_gc,
            completion_list=conditions,
            textvariable=self.var_condition,
            width=20,
        )
        self.ent_cond.pack(side=tk.LEFT, padx=5)
        r += 1

        # --- Action Bar ---
        action_bar = ttk.Frame(self, padding=20)
        action_bar.pack(fill=tk.X, side=tk.BOTTOM)

        ttk.Button(
            action_bar,
            text="SAVE ENTRY (Enter)",
            command=self._save_entry,
            style="Accent.TButton",
            width=20,
        ).pack(side=tk.RIGHT)
        ttk.Checkbutton(
            action_bar,
            text="Print Label after Save",
            variable=self.var_print_after_save,
        ).pack(side=tk.RIGHT, padx=15)

        self.lbl_status = ttk.Label(action_bar, text="Ready", foreground="gray")
        self.lbl_status.pack(side=tk.LEFT)

        # Configure Grid Weights
        self.form_frame.columnconfigure(1, weight=1)

        # --- BINDINGS ---
        self.field_order: list[tuple[tk.Widget, tk.BooleanVar | None]] = [
            (self.ent_imei, None),
            (self.ent_model, None),
            (self.ent_ram_rom, None),
            (self.cb_col, self.var_lock_color),
            (self.ent_price, None),
            (self.ent_supplier, self.var_lock_supplier),
            (self.cb_grade, self.var_lock_grade),
            (self.ent_cond, None),
        ]

        for i, (widget, lock_var) in enumerate(self.field_order):
            if widget == self.ent_imei:
                pass  # Already bound to _on_imei_enter
            elif widget == self.ent_cond:
                widget.bind("<Return>", lambda e: self._save_entry())
            else:
                widget.bind("<Return>", lambda e, idx=i: self._smart_focus_next(idx))

            widget.bind(
                "<Down>", lambda e, idx=i: self._smart_focus_next(idx, check_lock=False)
            )
            widget.bind("<Up>", lambda e, idx=i: self._focus_prev(idx))

        # Initialize View based on default mode
        self._toggle_batch_mode()

    # -- helpers -------------------------------------------------------------

    def _config_list(self, key: str, fallback: list[str]) -> list[str]:
        """Return a pick-list from config, falling back to *fallback*."""
        value = self.app["config"].get(key)
        if isinstance(value, list) and value:
            return value
        return list(fallback)

    def _get_unique_column_values(self, column: str) -> list[str]:
        """Return unique non-null values for *column* from the inventory."""
        df = self.app["inventory"].get_inventory()
        if df.empty or column not in df.columns:
            return []
        return df[column].dropna().unique().tolist()

    def _smart_focus_next(self, current_idx: int, check_lock: bool = True) -> None:
        """Move focus to the next unlocked field in the form."""
        next_idx = current_idx + 1
        if next_idx >= len(self.field_order):
            if current_idx == len(self.field_order) - 1:
                self._save_entry()
            return

        widget, lock_var = self.field_order[next_idx]

        if check_lock and lock_var is not None and lock_var.get():
            self._smart_focus_next(next_idx, check_lock=True)
        else:
            widget.focus_set()

    def _focus_prev(self, current_idx: int) -> None:
        """Move focus to the previous field in the form."""
        if current_idx > 0:
            widget, _ = self.field_order[current_idx - 1]
            widget.focus_set()

    # -- file management -----------------------------------------------------

    def _refresh_files(self) -> None:
        """Populate the target-file combobox from configured mappings."""
        mappings = self.app["config"].mappings
        raw_paths = list(mappings.keys())

        # Build display map from path basenames
        self.file_display_map = {}
        for path in raw_paths:
            display_name = Path(path).name
            # De-duplicate: append parent if basename collides
            if display_name in self.file_display_map:
                display_name = f"{Path(path).parent.name}/{Path(path).name}"
            self.file_display_map[display_name] = path

        self.file_options = sorted(self.file_display_map.keys())
        self.combo_file["values"] = self.file_options

        current_display = self.target_file_key.get()
        if self.file_options:
            if not current_display or current_display not in self.file_options:
                self.combo_file.current(0)
        else:
            self.combo_file.set("No Configured Files")

    # -- batch mode ----------------------------------------------------------

    def _toggle_batch_mode(self) -> None:
        """Switch between single-entry and batch-scan UI."""
        if self.var_batch_mode.get():
            self.form_frame.pack_forget()
            self.batch_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
            self._setup_batch_ui()
            self.ent_batch_scan.focus_set()
        else:
            self.batch_container.pack_forget()
            self.form_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
            self.ent_imei.focus_set()

    def _setup_batch_ui(self) -> None:
        """Build the batch-scan interface."""
        for w in self.batch_container.winfo_children():
            w.destroy()

        f_mid = ttk.Frame(self.batch_container)
        f_mid.place(relx=0.5, rely=0.4, anchor=tk.CENTER)

        ttk.Label(
            f_mid,
            text="BATCH SCAN MODE",
            font=("Segoe UI", 18, "bold"),
            foreground="#007acc",
        ).pack(pady=10)

        self.var_batch_scan = tk.StringVar()
        self.ent_batch_scan = ttk.Entry(
            f_mid,
            textvariable=self.var_batch_scan,
            width=40,
            font=("Consolas", 16),
        )
        self.ent_batch_scan.pack(pady=10)
        self.ent_batch_scan.bind("<Return>", self._on_imei_enter)

        self.lbl_queue_count = ttk.Label(
            f_mid,
            text="Items in Queue: 0",
            font=("Segoe UI", 12),
        )
        self.lbl_queue_count.pack(pady=5)

        f_log = ttk.Frame(f_mid)
        f_log.pack(fill=tk.BOTH, expand=True, pady=10)

        self.list_batch_log = tk.Listbox(
            f_log,
            height=8,
            width=50,
            font=("Consolas", 10),
            bg="#f8f9fa",
        )
        sb = ttk.Scrollbar(f_log, orient="vertical", command=self.list_batch_log.yview)
        self.list_batch_log.configure(yscrollcommand=sb.set)

        self.list_batch_log.pack(side=tk.LEFT, fill=tk.BOTH)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        self.list_batch_log.bind("<Delete>", self._remove_batch_item)
        self.list_batch_log.bind("<Button-3>", self._show_batch_context_menu)

        ttk.Label(
            f_mid,
            text="Scan IMEIs one by one.\nPress Ctrl+Enter when finished to start editing details.",
            justify=tk.CENTER,
            foreground="gray",
        ).pack(pady=20)

    def _show_batch_context_menu(self, event: tk.Event) -> None:
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(
            label="Remove Item",
            command=lambda: self._remove_batch_item(event),
        )
        menu.post(event.x_root, event.y_root)

    def _remove_batch_item(self, event: tk.Event | None = None) -> None:
        sel = self.list_batch_log.curselection()
        if not sel:
            return

        idx = sel[0]
        self.list_batch_log.delete(idx)

        queue_idx = len(self.batch_queue) - 1 - idx
        if 0 <= queue_idx < len(self.batch_queue):
            self.batch_queue.pop(queue_idx)

        self.lbl_queue_count.config(text=f"Items in Queue: {len(self.batch_queue)}")

    def _add_batch_to_queue(self, imei: str) -> None:
        if imei in self.batch_queue:
            return  # Avoid duplicates

        self.batch_queue.append(imei)
        self.lbl_queue_count.config(text=f"Items in Queue: {len(self.batch_queue)}")

        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self.list_batch_log.insert(0, f"[{ts}] Scanned: {imei}")

        def run() -> None:
            data = self.scraper.fetch_details(imei)
            self.fetched_data[imei] = data

        threading.Thread(target=run, daemon=True).start()

    def _focus_batch_start(self) -> None:
        """Switch from batch mode to single-entry and load first item."""
        if not self.var_batch_mode.get():
            return

        if not self.batch_queue:
            messagebox.showinfo("Empty Queue", "Please scan at least one IMEI first.")
            return

        self.var_batch_mode.set(False)
        self._toggle_batch_mode()
        self._load_next_from_batch()

    def _load_next_from_batch(self) -> None:
        """Pop the next IMEI from the batch queue and populate the form."""
        if not self.batch_queue:
            self.var_imei.set("")
            self.var_model.set("")
            self.lbl_status.config(text="Batch Completed!", foreground="green")
            self.ent_imei.focus_set()
            return

        imei = self.batch_queue.pop(0)
        self.var_imei.set(imei)

        data = self.fetched_data.get(imei)
        if data and not data.get("error") and data.get("name"):
            self.var_model.set(data["name"])
            self.lbl_status.config(
                text=f"Loaded from Batch (Remaining: {len(self.batch_queue)})",
                foreground="blue",
            )
        else:
            self.var_model.set("")
            self.lbl_status.config(
                text=f"Loading next... (Remaining: {len(self.batch_queue)})",
                foreground="blue",
            )
            self._fetch_info_bg(imei)

        self.ent_model.config(state="normal")
        self.ent_model.focus_set()
        self.ent_model.select_range(0, tk.END)

    # -- model entry ---------------------------------------------------------

    def _toggle_model_entry(self) -> None:
        if self.var_manual_model.get():
            self.ent_model.config(state="normal")
        else:
            self.ent_model.config(state="readonly")

    # -- IMEI handling -------------------------------------------------------

    def _on_imei_enter(self, event: tk.Event) -> None:
        is_batch = self.var_batch_mode.get()
        if is_batch:
            val = self.var_batch_scan.get().strip()
            self.var_batch_scan.set("")
        else:
            val = self.var_imei.get().strip()

        if not val:
            return

        # Immediate duplicate check
        df = self.app["inventory"].get_inventory()
        if not df.empty:
            mask = df["imei"].astype(str).str.contains(val, na=False, case=False)
            if mask.any():
                existing_model = df[mask].iloc[0].get("model", "Unknown")
                self.lbl_status.config(
                    text=f"ALREADY IN STOCK: {existing_model}",
                    foreground="red",
                )
            else:
                self.lbl_status.config(text="Ready", foreground="gray")

        is_real_imei = val.isdigit()

        if not is_batch and not is_real_imei:
            self.var_manual_model.set(True)
            self._toggle_model_entry()
            self.ent_model.focus_set()
            return

        if is_batch:
            self._add_batch_to_queue(val)
            self.ent_batch_scan.focus_set()
        else:
            self._smart_focus_next(0)
            if self.var_auto_fetch.get() and not self.var_manual_model.get():
                self._fetch_info_bg(val)

    # -- scraper -------------------------------------------------------------

    def _fetch_info_bg(
        self, imei: str, batch_row: dict[str, Any] | None = None
    ) -> None:
        if batch_row is None:
            self.lbl_status.config(text="Fetching info...", foreground="blue")

        def run() -> None:
            data = self.scraper.fetch_details(imei)
            self.after(0, lambda: self._on_fetch_complete(data, batch_row))

        threading.Thread(target=run, daemon=True).start()

    def _on_fetch_complete(
        self,
        data: dict[str, Any],
        batch_row: dict[str, Any] | None,
    ) -> None:
        name = data.get("name", "")

        if batch_row is not None:
            v_model = batch_row["vars"][1]
            v_model.set(name if name else "Not Found")
        else:
            if data.get("error"):
                self.lbl_status.config(
                    text=f"Fetch Error: {data['error']}",
                    foreground="red",
                )
            elif name:
                self.var_model.set(name)
                self.lbl_status.config(text="Details Fetched!", foreground="green")
            else:
                self.lbl_status.config(text="Model not found.", foreground="orange")

    # -- file creation -------------------------------------------------------

    def _create_new_file(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Create New Inventory File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialdir=self.app["config"].get("output_folder", "."),
        )
        if not path:
            return

        if os.path.exists(path):
            messagebox.showerror("Error", "File already exists!")
            return

        try:
            headers = [
                "S.NO",
                "SUPPLIER",
                "MODEL NAME",
                "RAM/ROM",
                "IMEI NO",
                "RATE",
                "STATUS",
                "BUYERS",
                "COLOR",
                "GRADE",
                "CONDITION",
            ]
            df = pd.DataFrame(columns=headers)
            df.to_excel(path, index=False)

            mapping = {
                "SUPPLIER": "supplier",
                "MODEL NAME": "model",
                "RAM/ROM": "ram_rom",
                "IMEI NO": "imei",
                "RATE": "price",
                "STATUS": "status",
                "BUYERS": "buyer",
                "COLOR": "color",
                "GRADE": "grade",
                "CONDITION": "condition",
            }
            save_data = {
                "file_path": path,
                "mapping": mapping,
                "supplier": "",
                "sheet_name": "Sheet1",
            }
            self.app["config"].set_file_mapping(path, save_data)

            self._refresh_files()

            display_name = next(
                (k for k, v in self.file_display_map.items() if v == path),
                None,
            )
            if display_name:
                self.target_file_key.set(display_name)

            messagebox.showinfo("Success", f"Created {os.path.basename(path)}")

        except Exception as exc:
            messagebox.showerror("Error", f"Failed to create file: {exc}")

    # -- save ----------------------------------------------------------------

    def _save_entry(self) -> None:
        imei = self.var_imei.get().strip()
        model = self.var_model.get().strip()

        target_display = self.target_file_key.get()
        target = self.file_display_map.get(target_display)

        if not imei or not model:
            messagebox.showwarning("Missing Data", "IMEI and Model are required.")
            if not imei:
                self.ent_imei.focus_set()
            else:
                self.ent_model.focus_set()
            return

        if not target or target_display == "No Configured Files":
            messagebox.showwarning("Target File", "Please select a target Excel file.")
            return

        # --- Duplicate IMEI check ---
        df = self.app["inventory"].get_inventory()
        if not df.empty and imei:
            mask = df["imei"].astype(str).str.contains(imei, na=False, case=False)
            if mask.any():
                existing = df[mask].iloc[0]
                msg = (
                    f"DUPLICATE IMEI DETECTED!\n\n"
                    f"IMEI: {imei}\n"
                    f"Model: {existing.get('model', 'Unknown')}\n"
                    f"Status: {existing.get('status', 'Unknown')}\n"
                    f"File: {os.path.basename(str(existing.get('source_file', 'Unknown')))}\n\n"
                    f"This item already exists in your inventory.\n"
                    f"Do you want to FORCE SAVE this as a new entry?"
                )
                if not messagebox.askyesno("Duplicate IMEI", msg):
                    self.lbl_status.config(
                        text="Save cancelled: Duplicate IMEI",
                        foreground="orange",
                    )
                    self.ent_imei.focus_set()
                    self.ent_imei.select_range(0, tk.END)
                    return

        # Prepare data
        price = 0.0
        try:
            p_str = self.var_price.get().strip()
            if p_str:
                price = float(p_str)
        except (ValueError, TypeError):
            pass

        new_data = {
            "imei": imei,
            "model": model,
            "ram_rom": self.var_ram_rom.get().strip(),
            "color": self.var_color.get().strip(),
            "price_original": price,
            "supplier": self.var_supplier.get().strip(),
            "grade": self.var_grade.get().strip(),
            "condition": self.var_condition.get().strip(),
            "status": "IN",
            "source_file": target,
        }

        # 1. Create ID via database
        uid = self.app["db"].get_or_create_id(
            imei=imei,
            model=model,
            ram_rom=new_data["ram_rom"],
            supplier=new_data["supplier"],
            source_file=target,
            color=new_data["color"],
            price_original=new_data["price_original"],
            grade=new_data["grade"],
            condition=new_data["condition"],
        )
        new_data["unique_id"] = uid

        # 2. Append to Excel
        success = self._append_to_excel(target, new_data)

        if success:
            self.lbl_status.config(text=f"Saved ID: {uid}", foreground="green")

            if self.var_print_after_save.get():
                self._print_label(new_data)

            # 4. Clear fields (respect locks)
            self.var_imei.set("")
            self.var_model.set("")
            self.var_ram_rom.set("")
            self.var_price.set("")

            if not self.var_lock_color.get():
                self.var_color.set("")
            if not self.var_lock_supplier.get():
                self.var_supplier.set("")
            if not self.var_lock_grade.get():
                self.var_grade.set("")
                self.var_condition.set("")

            # --- Batch continuation ---
            if self.batch_queue:
                self._load_next_from_batch()
            else:
                self.ent_imei.focus_set()
        else:
            messagebox.showerror("Error", "Failed to save to Excel file.")

    def _append_to_excel(self, file_key: str, data: dict[str, Any]) -> bool:
        path = file_key
        sheet: str | int = 0
        if "::" in file_key:
            parts = file_key.split("::")
            path = parts[0]
            sheet = parts[1]

        mapping_data = self.app["config"].get_file_mapping(file_key)
        if not mapping_data:
            return False

        mapping = mapping_data.get("mapping", {})
        field_to_col = {v: k for k, v in mapping.items()}

        row_dict: dict[str, Any] = {}
        for field, val in data.items():
            if field == "price_original":
                field = "price"
            if field in field_to_col:
                row_dict[field_to_col[field]] = val

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Border, Font, Side

            wb = load_workbook(path)

            ws = None
            if isinstance(sheet, str) and sheet in wb.sheetnames:
                ws = wb[sheet]
            else:
                ws = wb.active

            col_map: dict[str, int] = {}
            for cell in ws[1]:
                if cell.value:
                    col_map[str(cell.value).strip()] = cell.column

            next_row = ws.max_row + 1

            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            font = Font(name="Times New Roman", size=11, bold=True)
            align = Alignment(horizontal="center", vertical="center")

            for col_name, val in row_dict.items():
                if col_name in col_map:
                    if isinstance(val, str):
                        val = val.upper()
                    cell = ws.cell(row=next_row, column=col_map[col_name], value=val)
                    cell.font = font
                    cell.border = border
                    cell.alignment = align

            wb.save(path)
            return True
        except Exception as exc:
            logger.error("Append Error: %s", exc)
            return False

    # -- printing ------------------------------------------------------------

    def _print_label(self, item: dict[str, Any]) -> None:
        """Print a label for *item* using the configured ZPL template."""
        self.app["printer"].print_label_zpl(item)
