import logging
import pandas as pd
import os
import re
import queue
import threading
import datetime
import shutil
import time
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

from .config import ConfigManager
from .database import SQLiteDatabase
from .utils import backup_excel_file
from .constants import (
    STATUS_IN,
    STATUS_OUT,
    STATUS_RETURN,
    ACTION_STATUS_CHANGE,
    ACTION_DATA_UPDATE,
    ACTION_MERGE,
    ACTION_RELOAD,
    ACTION_REDIRECT,
    ACTION_ITEM_UPDATE,
    ACTION_RESOLVE_CONFLICT,
    FIELD_IMEI,
    FIELD_MODEL,
    FIELD_PRICE,
    FIELD_STATUS,
    FIELD_BUYER,
    FIELD_BUYER_CONTACT,
    FIELD_UNIQUE_ID,
    FIELD_SOURCE_FILE,
    FIELD_NOTES,
    FIELD_COLOR,
    FIELD_RAM_ROM,
    FIELD_PRICE_ORIGINAL,
    FIELD_GRADE,
    FIELD_CONDITION,
)


class InventoryManager:
    """Core inventory manager — loads, normalises, merges, and persists stock data."""

    def __init__(
        self,
        config_manager: ConfigManager,
        db: SQLiteDatabase,
        activity_logger: Any = None,
    ) -> None:
        self.config_manager = config_manager
        self.db = db
        self.activity_logger = activity_logger

        self.inventory_df: pd.DataFrame = pd.DataFrame()
        self._df_lock = threading.RLock()
        self.file_status: dict[str, str] = {}
        self.conflicts: list[dict[str, Any]] = []

        # Background write queue
        self.write_queue: queue.Queue = queue.Queue()
        self._writer_thread: threading.Thread | None = None
        self._start_worker()

    # ------------------------------------------------------------------
    # Background worker
    # ------------------------------------------------------------------

    def _start_worker(self) -> None:
        """Spawn a daemon thread that drains *write_queue* sequentially."""

        def worker() -> None:
            while True:
                task = self.write_queue.get()
                if task is None:
                    break
                try:
                    func, args = task
                    func(*args)
                except Exception as exc:
                    logger.error("Background Worker Error: %s", exc)
                finally:
                    self.write_queue.task_done()

        self._writer_thread = threading.Thread(
            target=worker, daemon=True, name="ExcelWriterThread"
        )
        self._writer_thread.start()

    def shutdown(self) -> None:
        """Drain pending writes, send sentinel, and join the worker thread."""
        try:
            self.write_queue.join()
        except Exception:
            pass

        self.write_queue.put(None)

        if self._writer_thread is not None:
            self._writer_thread.join(timeout=10)

    # ------------------------------------------------------------------
    # File loading
    # ------------------------------------------------------------------

    def load_file(self, file_path: str) -> tuple[pd.DataFrame | None, str]:
        """Public entry-point — resolve mapping then delegate."""
        mapping_data = self.config_manager.get_file_mapping(file_path)
        return self._load_file_internal(file_path, mapping_data)

    def _load_file_internal(
        self, file_path: str, mapping_data: dict | None
    ) -> tuple[pd.DataFrame | None, str]:
        """Read Excel or CSV, normalise, return (DataFrame, status)."""
        if not mapping_data:
            return None, "MAPPING_REQUIRED"

        try:
            if file_path.endswith(".csv"):
                df = pd.read_csv(file_path)
            else:
                sheet_name = mapping_data.get("sheet_name", 0)
                # Convert numeric strings to int so pandas treats as index, not name
                if isinstance(sheet_name, str) and sheet_name.isdigit():
                    sheet_name = int(sheet_name)
                if not sheet_name:
                    sheet_name = 0

                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                except (ValueError, IndexError, KeyError) as exc:
                    # If sheet_name fails, try loading the first sheet as fallback
                    try:
                        df = pd.read_excel(file_path, sheet_name=0)
                    except Exception as exc2:
                        return None, f"SHEET_ERROR: {exc2}"

            if isinstance(df, pd.Series):
                df = df.to_frame().T

            normalised = self._normalize_data(df, mapping_data, file_path)
            return normalised, "SUCCESS"

        except Exception as exc:
            return None, str(exc)

    # ------------------------------------------------------------------
    # Normalisation — the heavy lifter
    # ------------------------------------------------------------------

    def _normalize_data(
        self, df: pd.DataFrame, mapping_data: dict, file_path: str
    ) -> pd.DataFrame:
        """Convert raw Excel/CSV data into the canonical schema."""

        mapping: dict[str, str] = mapping_data.get("mapping") or {}
        file_supplier: str = mapping_data.get("supplier", "")

        # ---- pure helpers ------------------------------------------------

        def norm_status(val: Any) -> str:
            s = str(val).upper().strip()
            if s in (STATUS_OUT, "SOLD", "SALE"):
                return STATUS_OUT
            if s in (STATUS_RETURN, "RETURN", "RET"):
                return STATUS_RETURN
            return STATUS_IN

        def get_col(canonical_name: str) -> pd.Series | None:
            """Return the raw Series that maps to *canonical_name*, or None."""
            for raw_col, can_name in mapping.items():
                if can_name == canonical_name and raw_col in df.columns:
                    return df[raw_col].astype(object)
            return None

        def clean_imei(val: Any) -> str:
            if pd.isna(val):
                return ""
            s = str(val).strip()
            digits = re.findall(r"\d{14,16}", s)
            if len(digits) > 1:
                return " / ".join(sorted(set(digits)))
            if len(digits) == 1:
                return digits[0]
            return s

        # ---- build canonical DataFrame -----------------------------------

        canonical = pd.DataFrame()

        # IMEI
        col_imei = get_col(FIELD_IMEI)
        canonical[FIELD_IMEI] = (
            col_imei.apply(clean_imei) if col_imei is not None else ""
        )

        # Model
        col_model = get_col(FIELD_MODEL)
        canonical[FIELD_MODEL] = (
            col_model.fillna("Unknown Model").astype(str)
            if col_model is not None
            else "Unknown Model"
        )

        # Brand
        col_brand = get_col("brand")
        if col_brand is not None:
            canonical["brand"] = col_brand.fillna("").astype(str).str.upper()
        else:
            canonical["brand"] = canonical[FIELD_MODEL].apply(
                lambda x: str(x).split()[0].upper() if str(x).split() else "UNKNOWN"
            )

        # Price
        col_price = get_col(FIELD_PRICE)
        if col_price is not None:
            raw_price = pd.to_numeric(col_price, errors="coerce").fillna(0.0)
        else:
            raw_price = 0.0

        canonical[FIELD_PRICE_ORIGINAL] = raw_price

        try:
            markup = float(self.config_manager.get("price_markup_percent", 0.0))
        except (ValueError, TypeError):
            markup = 0.0

        if markup > 0:
            price_with_markup = raw_price * (1 + markup / 100.0)
            canonical[FIELD_PRICE] = price_with_markup.apply(
                lambda x: round(x / 100) * 100 if x > 0 else x
            )
        else:
            canonical[FIELD_PRICE] = raw_price

        # RAM / ROM
        col_ram_rom = get_col(FIELD_RAM_ROM)
        if col_ram_rom is not None:
            canonical[FIELD_RAM_ROM] = col_ram_rom.fillna("").astype(str)
        else:
            col_ram = get_col("ram")
            col_rom = get_col("rom")
            if col_ram is not None and col_rom is not None:
                canonical[FIELD_RAM_ROM] = (
                    col_ram.fillna("").astype(str)
                    + " / "
                    + col_rom.fillna("").astype(str)
                )
            elif col_ram is not None:
                canonical[FIELD_RAM_ROM] = col_ram.fillna("").astype(str)
            else:
                # Fallback: try common Excel column name variations
                ram_rom_found = False
                for raw_col in df.columns:
                    raw_lower = str(raw_col).strip().lower()
                    if raw_lower in (
                        "ram/rom",
                        "ram / rom",
                        "ram_rom",
                        "storage",
                        "memory",
                        "internal storage",
                        "internal",
                        "ram + rom",
                        "ram&rom",
                        "ram & rom",
                    ):
                        canonical[FIELD_RAM_ROM] = df[raw_col].fillna("").astype(str)
                        ram_rom_found = True
                        break
                if not ram_rom_found:
                    canonical[FIELD_RAM_ROM] = ""

        # Supplier
        col_supplier = get_col("supplier")
        if col_supplier is not None:
            canonical["supplier"] = col_supplier.fillna(file_supplier).astype(str)
        else:
            canonical["supplier"] = file_supplier

        # Status
        col_status = get_col(FIELD_STATUS)
        if col_status is not None:
            canonical[FIELD_STATUS] = col_status.apply(norm_status)
        else:
            canonical[FIELD_STATUS] = STATUS_IN

        # Color
        col_color = get_col(FIELD_COLOR)
        canonical[FIELD_COLOR] = (
            col_color.fillna("").astype(str) if col_color is not None else ""
        )

        # Buyer
        col_buyer = get_col(FIELD_BUYER)
        canonical[FIELD_BUYER] = (
            col_buyer.fillna("").astype(str) if col_buyer is not None else ""
        )

        col_contact = get_col(FIELD_BUYER_CONTACT)
        canonical[FIELD_BUYER_CONTACT] = (
            col_contact.fillna("").astype(str) if col_contact is not None else ""
        )

        # Grade / Condition
        col_grade = get_col(FIELD_GRADE)
        canonical[FIELD_GRADE] = (
            col_grade.fillna("").astype(str) if col_grade is not None else ""
        )

        col_condition = get_col(FIELD_CONDITION)
        canonical[FIELD_CONDITION] = (
            col_condition.fillna("").astype(str) if col_condition is not None else ""
        )

        # Metadata columns
        canonical[FIELD_SOURCE_FILE] = str(file_path)
        now_ts = datetime.datetime.now()
        canonical["last_updated"] = now_ts
        canonical["date_added"] = now_ts

        # ---- ID generation via database ----------------------------------

        canonical[FIELD_UNIQUE_ID] = canonical.apply(
            lambda row: self.db.get_or_create_id(
                imei=row[FIELD_IMEI],
                model=row[FIELD_MODEL],
                ram_rom=row[FIELD_RAM_ROM],
                supplier=row["supplier"],
                source_file=file_path,
                color=row.get(FIELD_COLOR, ""),
                price_original=row.get(FIELD_PRICE_ORIGINAL, 0.0),
                grade=row.get(FIELD_GRADE, ""),
                condition=row.get(FIELD_CONDITION, ""),
            ),
            axis=1,
        )

        # ---- Merge persistent metadata from DB ---------------------------

        unique_ids = canonical[FIELD_UNIQUE_ID].unique().tolist()
        int_ids = [int(uid) for uid in unique_ids]
        items_data = self.db.get_items_by_ids(int_ids)
        item_map: dict[int, dict[str, Any]] = {item["id"]: item for item in items_data}

        def apply_overrides(row: pd.Series) -> pd.Series:
            uid = int(row[FIELD_UNIQUE_ID])
            item_data = item_map.get(uid)

            if not item_data:
                return row

            # Restore added_date
            if item_data.get("added_date"):
                try:
                    row["date_added"] = datetime.datetime.fromisoformat(
                        item_data["added_date"]
                    )
                except (ValueError, TypeError):
                    pass

            # Override status from DB
            if item_data.get("status"):
                row[FIELD_STATUS] = norm_status(item_data["status"])

            # Restore sold_date
            if item_data.get("sold_date"):
                try:
                    row["date_sold"] = datetime.datetime.fromisoformat(
                        item_data["sold_date"]
                    )
                except (ValueError, TypeError):
                    pass

            # Override notes
            if item_data.get("notes"):
                row[FIELD_NOTES] = item_data["notes"]

            # Override color, grade, condition from stored values
            if item_data.get(FIELD_COLOR):
                row[FIELD_COLOR] = item_data[FIELD_COLOR]
            if item_data.get(FIELD_GRADE):
                row[FIELD_GRADE] = item_data[FIELD_GRADE]
            if item_data.get(FIELD_CONDITION):
                row[FIELD_CONDITION] = item_data[FIELD_CONDITION]

            # Price override
            if item_data.get("price_override") is not None:
                row[FIELD_PRICE_ORIGINAL] = float(item_data["price_override"])
                try:
                    m = float(self.config_manager.get("price_markup_percent", 0.0))
                except (ValueError, TypeError):
                    m = 0.0
                if m > 0:
                    raw_p = row[FIELD_PRICE_ORIGINAL] * (1 + m / 100.0)
                    row[FIELD_PRICE] = round(raw_p / 100) * 100 if raw_p > 0 else raw_p
                else:
                    row[FIELD_PRICE] = row[FIELD_PRICE_ORIGINAL]

            return row

        canonical = canonical.apply(apply_overrides, axis=1)

        # ---- Ensure all required columns exist ---------------------------

        required_cols = [
            FIELD_UNIQUE_ID,
            FIELD_IMEI,
            "brand",
            FIELD_MODEL,
            FIELD_RAM_ROM,
            FIELD_PRICE,
            FIELD_PRICE_ORIGINAL,
            "supplier",
            FIELD_SOURCE_FILE,
            "last_updated",
            "date_added",
            "date_sold",
            FIELD_STATUS,
            FIELD_COLOR,
            FIELD_NOTES,
            FIELD_BUYER,
            FIELD_BUYER_CONTACT,
            FIELD_GRADE,
            FIELD_CONDITION,
        ]
        for col in required_cols:
            if col not in canonical.columns:
                canonical[col] = (
                    0.0 if col in (FIELD_PRICE, FIELD_PRICE_ORIGINAL) else ""
                )

        return canonical

    # ------------------------------------------------------------------
    # Reload — merge all mapped files
    # ------------------------------------------------------------------

    def reload_all(self) -> pd.DataFrame:
        """Reload every mapped file, merge, filter hidden, detect conflicts."""
        self.conflicts = []
        all_frames: list[pd.DataFrame] = []
        mappings = self.config_manager.mappings

        for key, mapping_data in mappings.items():
            file_path = mapping_data.get("file_path", key)

            # Composite key fallback: "path::sheet" → "path"
            if "::" in key and not os.path.exists(key):
                parts = key.split("::")
                if os.path.exists(parts[0]):
                    file_path = parts[0]

            if os.path.exists(file_path):
                df, status = self._load_file_internal(file_path, mapping_data)
                if status == "SUCCESS" and df is not None:
                    df[FIELD_SOURCE_FILE] = key
                    all_frames.append(df)
                    self.file_status[key] = "OK"
                else:
                    self.file_status[key] = f"Error: {status}"
            else:
                self.file_status[key] = "Missing"

        if all_frames:
            full_df = pd.concat(all_frames, ignore_index=True)

            # Filter hidden items
            unique_ids = full_df[FIELD_UNIQUE_ID].astype(int).unique().tolist()
            items_data = self.db.get_items_by_ids(unique_ids)
            hidden_ids = {item["id"] for item in items_data if item.get("is_hidden")}
            if hidden_ids:
                full_df = full_df[~full_df[FIELD_UNIQUE_ID].isin(hidden_ids)]

            # Detect IMEI conflicts
            self._detect_conflicts(full_df)

            with self._df_lock:
                self.inventory_df = full_df
        else:
            with self._df_lock:
                self.inventory_df = pd.DataFrame(
                    columns=[
                        FIELD_UNIQUE_ID,
                        FIELD_IMEI,
                        "brand",
                        FIELD_MODEL,
                        FIELD_RAM_ROM,
                        FIELD_PRICE,
                        FIELD_PRICE_ORIGINAL,
                        "supplier",
                        FIELD_SOURCE_FILE,
                        "last_updated",
                        "date_added",
                        "date_sold",
                        FIELD_STATUS,
                        FIELD_COLOR,
                        FIELD_NOTES,
                        FIELD_BUYER,
                        FIELD_BUYER_CONTACT,
                        FIELD_GRADE,
                        FIELD_CONDITION,
                    ]
                )

        if self.activity_logger:
            self.activity_logger.log(
                ACTION_RELOAD,
                f"Loaded {len(self.inventory_df)} items from {len(mappings)} sources.",
            )

        return self.inventory_df

    def _detect_conflicts(self, df: pd.DataFrame) -> None:
        """Find duplicate real IMEIs and populate *self.conflicts*."""
        imei_df = df[df[FIELD_IMEI].str.len() > 5].copy()
        if imei_df.empty:
            return

        imei_df["imei_list"] = imei_df[FIELD_IMEI].str.split("/")
        exploded = imei_df.explode("imei_list")
        exploded["imei_list"] = exploded["imei_list"].str.strip()
        exploded = exploded[exploded["imei_list"].str.len() > 5]

        dupes = exploded[exploded.duplicated("imei_list", keep=False)]
        if dupes.empty:
            return

        for imei, group in dupes.groupby("imei_list"):
            self.conflicts.append(
                {
                    "imei": imei,
                    "unique_ids": group[FIELD_UNIQUE_ID].unique().tolist(),
                    "model": group.iloc[0][FIELD_MODEL],
                    "sources": group[FIELD_SOURCE_FILE].unique().tolist(),
                    "rows": group.drop_duplicates(FIELD_UNIQUE_ID).to_dict("records"),
                }
            )

    # ------------------------------------------------------------------
    # Conflict resolution
    # ------------------------------------------------------------------

    def resolve_conflict(
        self, keep_id: int, hide_ids: list[int], reason: str = ""
    ) -> bool:
        """Hide *hide_ids*, mark as merged into *keep_id*, then reload."""
        try:
            self.db.resolve_conflict(keep_id, hide_ids, reason)

            # Record keeper history
            self.db.add_history(
                keep_id,
                ACTION_RESOLVE_CONFLICT,
                f"Kept as primary. {len(hide_ids)} item(s) merged. Reason: {reason}",
            )

            self.reload_all()
            return True
        except Exception as exc:
            logger.error("Conflict resolution failed: %s", exc)
            return False

    # ------------------------------------------------------------------
    # Lookups
    # ------------------------------------------------------------------

    def get_item_by_id(
        self, unique_id: int | str, resolve_merged: bool = True
    ) -> tuple[dict[str, Any] | None, int | None]:
        """Return (item_dict, redirected_from_id) or (None, None)."""
        try:
            item_id = int(unique_id)
        except (ValueError, TypeError):
            return None, None

        target_id = item_id

        if resolve_merged:
            meta = self.db.get_metadata(item_id)
            if meta and meta.get("is_hidden"):
                merged_into = meta.get("merged_into")
                if merged_into is not None:
                    target_id = int(merged_into)

        with self._df_lock:
            mask = self.inventory_df[FIELD_UNIQUE_ID] == target_id
            if not mask.any():
                return None, None

            item = self.inventory_df[mask].iloc[0].to_dict()

            # Merge DB metadata (buyer, buyer_contact, notes, etc.)
            meta = self.db.get_metadata(target_id)
            if meta:
                for key in (FIELD_BUYER, FIELD_BUYER_CONTACT, FIELD_NOTES, "date_sold"):
                    if key in meta and meta[key] is not None:
                        item[key] = meta[key]

            redirected_from = item_id if target_id != item_id else None
            return item, redirected_from

    def get_inventory(self) -> pd.DataFrame:
        """Return a thread-safe copy of the current inventory."""
        with self._df_lock:
            return self.inventory_df.copy()

    # ------------------------------------------------------------------
    # Updates
    # ------------------------------------------------------------------

    def update_item_status(
        self, item_id: int | str, new_status: str, write_to_excel: bool = False
    ) -> bool:
        """Persist a status change, update memory, optionally queue Excel write."""
        try:
            numeric_id = int(item_id)
        except (ValueError, TypeError):
            return False

        # Redirect if merged
        meta = self.db.get_metadata(numeric_id)
        if meta and meta.get("is_hidden"):
            merged_into = meta.get("merged_into")
            if merged_into is not None:
                if self.activity_logger:
                    self.activity_logger.log(
                        ACTION_REDIRECT,
                        f"Status update for {item_id} redirected to {merged_into}",
                    )
                numeric_id = int(merged_into)

        with self._df_lock:
            mask = self.inventory_df[FIELD_UNIQUE_ID] == numeric_id
            old_status = "UNKNOWN"
            item_model = ""
            if mask.any():
                row = self.inventory_df[mask].iloc[0]
                old_status = row[FIELD_STATUS]
                item_model = row[FIELD_MODEL]

            # Build metadata updates
            updates: dict[str, Any] = {FIELD_STATUS: new_status}
            if new_status == STATUS_OUT:
                updates["sold_date"] = datetime.datetime.now().isoformat()
            elif new_status == STATUS_IN:
                updates["sold_date"] = ""

            self.db.update_metadata(numeric_id, **updates)
            self.db.add_history(
                numeric_id,
                ACTION_STATUS_CHANGE,
                f"Moved from {old_status} to {new_status}",
            )

            if self.activity_logger:
                self.activity_logger.log(
                    ACTION_STATUS_CHANGE,
                    f"Item {numeric_id} ({item_model}) marked as {new_status}",
                )

            # Update in-memory DataFrame
            if mask.any():
                self.inventory_df.loc[mask, FIELD_STATUS] = new_status

                if write_to_excel:
                    try:
                        row_snapshot = self.inventory_df[mask].iloc[0].to_dict()
                        self.write_queue.put(
                            (
                                self._write_excel_generic,
                                (row_snapshot, {FIELD_STATUS: new_status}),
                            )
                        )
                    except Exception as exc:
                        logger.error("Queue Error: %s", exc)

        return True

    def update_item_data(self, item_id: int | str, updates: dict[str, Any]) -> bool:
        """Persist arbitrary field updates, update memory, queue Excel write."""
        try:
            numeric_id = int(item_id)
        except (ValueError, TypeError):
            return False

        # Redirect if merged
        meta = self.db.get_metadata(numeric_id)
        if meta and meta.get("is_hidden"):
            merged_into = meta.get("merged_into")
            if merged_into is not None:
                numeric_id = int(merged_into)

        with self._df_lock:
            mask = self.inventory_df[FIELD_UNIQUE_ID] == numeric_id
            if not mask.any():
                return False

            item_model = self.inventory_df.loc[mask, FIELD_MODEL].values[0]

            self.db.update_metadata(numeric_id, **updates)

            log_details = ", ".join(f"{k}={v}" for k, v in updates.items())
            self.db.add_history(numeric_id, ACTION_DATA_UPDATE, log_details)

            if self.activity_logger:
                self.activity_logger.log(
                    ACTION_ITEM_UPDATE,
                    f"Updated {numeric_id} ({item_model}): {log_details}",
                )

            # Update in-memory DataFrame
            for key, value in updates.items():
                if key in self.inventory_df.columns:
                    self.inventory_df.loc[mask, key] = value

            # Queue async Excel write
            try:
                row_snapshot = self.inventory_df[mask].iloc[0].to_dict()
                self.write_queue.put(
                    (self._write_excel_generic, (row_snapshot, dict(updates)))
                )
            except Exception as exc:
                logger.error("Queue error: %s", exc)

        return True

    # ------------------------------------------------------------------
    # Async Excel writer — runs on background thread
    # ------------------------------------------------------------------

    def _write_excel_generic(
        self, row_data: dict[str, Any], updates: dict[str, Any]
    ) -> tuple[bool, str]:
        """Background worker: locate row in source Excel and apply updates."""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side

        # Resolve file path (handle composite keys "path::sheet")
        key = row_data.get(FIELD_SOURCE_FILE, "")
        file_path = key
        if "::" in key and not os.path.exists(key):
            file_path = key.split("::")[0]

        if not file_path or not os.path.exists(file_path):
            msg = f"Source file not found: {file_path}"
            logger.error("Write Error: %s", msg)
            return False, msg

        # Safety backup
        backup_path = backup_excel_file(file_path)
        if not backup_path:
            msg = "Failed to create backup, aborting write for safety."
            logger.error("Write Error: %s", msg)
            return False, msg

        # Resolve mapping
        mapping_data = self.config_manager.get_file_mapping(key)
        if not mapping_data:
            mapping_data = self.config_manager.get_file_mapping(file_path)
        if not mapping_data:
            msg = f"No mapping found for {file_path}"
            logger.error("Write Error: %s", msg)
            return False, msg

        mapping: dict[str, str] = mapping_data.get("mapping") or {}
        field_to_col = {v: k for k, v in mapping.items()}

        # Map internal fields → Excel headers
        default_headers: dict[str, str] = {
            FIELD_BUYER: "Buyer Name",
            FIELD_BUYER_CONTACT: "Buyer Contact",
            FIELD_NOTES: "Notes",
            FIELD_STATUS: "Status",
            FIELD_COLOR: "Color",
            FIELD_PRICE: "Selling Price",
            FIELD_GRADE: "Grade",
            FIELD_CONDITION: "Condition",
        }

        excel_updates: dict[str, Any] = {}
        for field, value in updates.items():
            header_name = field_to_col.get(field) or default_headers.get(field)
            if header_name:
                excel_updates[header_name] = value

        max_retries = 3
        retry_delay = 1.5

        for attempt in range(max_retries):
            try:
                wb = load_workbook(file_path)

                # Determine sheet
                sheet_name = mapping_data.get("sheet_name", 0)
                if "::" in key:
                    sheet_part = key.split("::")[1]
                    if sheet_part:
                        sheet_name = (
                            int(sheet_part) if sheet_part.isdigit() else sheet_part
                        )

                ws = None
                if isinstance(sheet_name, int):
                    try:
                        ws = wb.worksheets[sheet_name]
                    except IndexError:
                        pass
                elif isinstance(sheet_name, str):
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    elif sheet_name.isdigit():
                        try:
                            ws = wb.worksheets[int(sheet_name)]
                        except (IndexError, ValueError):
                            pass

                if ws is None:
                    ws = wb.active

                # Build column index from header row
                col_indices: dict[str, int] = {}
                max_col_idx = 0
                for cell in ws[1]:
                    val = str(cell.value).strip() if cell.value else ""
                    if val:
                        col_indices[val] = cell.column
                        if cell.column > max_col_idx:
                            max_col_idx = cell.column

                # Create missing columns
                for header_name in excel_updates:
                    if header_name not in col_indices:
                        max_col_idx += 1
                        ws.cell(row=1, column=max_col_idx).value = header_name
                        col_indices[header_name] = max_col_idx

                # Find matching row by IMEI
                target_imei = str(row_data.get(FIELD_IMEI, ""))
                imei_header = field_to_col.get(FIELD_IMEI)
                imei_col_idx = col_indices.get(imei_header) if imei_header else None

                row_found = False
                target_parts = {p.strip() for p in target_imei.split("/") if p.strip()}

                for row in ws.iter_rows(min_row=2):
                    if not imei_col_idx:
                        continue

                    cell_val = row[imei_col_idx - 1].value
                    if not cell_val:
                        continue

                    s_cell = str(cell_val).strip().replace(".0", "")
                    cell_parts = {p.strip() for p in s_cell.split("/") if p.strip()}

                    if not target_parts or not cell_parts:
                        continue

                    # Exact match or dual-IMEI overlap
                    if target_parts == cell_parts or (target_parts & cell_parts):
                        row_found = True

                        for col_name, new_val in excel_updates.items():
                            if col_name not in col_indices:
                                continue
                            cell = ws.cell(row=row[0].row, column=col_indices[col_name])
                            if isinstance(new_val, str):
                                new_val = new_val.upper()
                            cell.value = new_val

                            thin = Side(border_style="thin", color="000000")
                            cell.border = Border(
                                top=thin, left=thin, right=thin, bottom=thin
                            )
                            cell.font = Font(name="Times New Roman", size=11, bold=True)
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                        break

                if not row_found:
                    msg = f"Row not found for IMEI: {target_imei}"
                    logger.warning("Warning: %s", msg)
                    return False, msg

                # Staged write: .tmp → atomic replace
                temp_path = f"{file_path}.tmp"
                wb.save(temp_path)

                if os.path.exists(temp_path):
                    os.replace(temp_path, file_path)
                    return True, "Success"
                return False, "Failed to write temp file"

            except PermissionError:
                if attempt < max_retries - 1:
                    logger.warning(
                        "Write Attempt %d failed: "
                        "File open in Excel. Retrying in %ss...",
                        attempt + 1,
                        retry_delay,
                    )
                    time.sleep(retry_delay)
                else:
                    return False, "File is open in Excel. Please close it."

            except Exception as exc:
                return False, f"Excel Write Error: {exc}"

        return False, "Max retries exceeded"
